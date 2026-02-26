#!/usr/bin/env python3
"""Inspect Sensores.xlsx structure"""
from openpyxl import load_workbook

excel_file = 'docs/Sensores.xlsx'
wb = load_workbook(excel_file)

print(f'[*] Sheets: {wb.sheetnames}')
print()

ws = wb[wb.sheetnames[0]]
print(f'[*] Active sheet: {ws.title}')
print(f'[*] Max row: {ws.max_row}')
print(f'[*] Max column: {ws.max_column}')
print()

print('[*] Header row (columns A-E):')
for col in range(1, min(6, ws.max_column + 1)):
    cell = ws.cell(1, col)
    col_letter = chr(64 + col)
    print(f'    Column {col_letter}: {cell.value}')

print()
print('[*] First 3 data rows:')
for row in range(2, min(5, ws.max_row + 1)):
    values = []
    for col in range(1, min(5, ws.max_column + 1)):
        cell = ws.cell(row, col)
        # Get value, handling formulas
        if hasattr(cell, 'value') and cell.value:
            val = str(cell.value)[:35]
        else:
            val = 'None'
        values.append(val)
    print(f'    Row {row}: {" || ".join(values)}')

print()
print(f'[*] Total data rows: {ws.max_row - 1}')
