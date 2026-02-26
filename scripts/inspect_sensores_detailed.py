#!/usr/bin/env python3
"""Inspect Sensores.xlsx with formula calculation"""
from openpyxl import load_workbook

print("[*] Reading Sensores.xlsx...")
excel_file = 'docs/Sensores.xlsx'

# Load with data_only to get calculated values
wb = load_workbook(excel_file, data_only=True)
ws = wb[wb.sheetnames[0]]

print(f'[*] Sheet: {ws.title}')
print(f'[*] Total rows: {ws.max_row}')
print(f'[*] Total columns: {ws.max_column}')
print()

print('[*] Column Headers:')
headers = []
for col in range(1, ws.max_column + 1):
    cell = ws.cell(1, col)
    headers.append(cell.value)
    col_letter = chr(64 + col)
    print(f'    {col_letter}: {cell.value}')

print()
print('[*] First 5 data rows:')
print('-' * 100)
for row in range(2, min(7, ws.max_row + 1)):
    row_data = {}
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row, col)
        row_data[header] = cell.value
    
    print(f'Row {row}:')
    for header, value in row_data.items():
        print(f'  {header:15} = {str(value)[:50] if value else "None"}')
    print()

print(f'[*] Total sensores: {ws.max_row - 1}')
