#!/usr/bin/env python3
"""
Script para limpar o banco de dados, mantendo apenas os sensores da planilha Sensores.xlsx.

Este script:
1. Lê a planilha docs/Sensores.xlsx
2. Extrai a lista de sensores
3. Deleta do banco todos os sensores que NÃO estão na planilha
4. Mantém APENAS os 9.983 sensores da planilha

Uso:
    python scripts/cleanup_db_keep_sensores_xlsx.py [--verify-only]
"""

import sys
import os
from pathlib import Path

# Adicionar o backend ao path
sys.path.insert(0, str(Path(__file__).parent.parent / "backend"))

from sqlalchemy import create_engine
from sqlalchemy.orm import sessionmaker
from openpyxl import load_workbook

from src.data.models import SensorConfig

def extract_sensors_from_xlsx():
    """Extrai lista de sensores da planilha Sensores.xlsx"""
    
    excel_file = Path(__file__).parent.parent / "docs" / "Sensores.xlsx"
    
    if not excel_file.exists():
        raise FileNotFoundError(f"Arquivo não encontrado: {excel_file}")
    
    print(f"\n[*] Lendo planilha: {excel_file}")
    
    wb = load_workbook(excel_file, data_only=True)
    ws = wb.active
    
    sensores = set()
    linhas = 0
    
    # Pular header (linha 1)
    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if row[0] is None:  # Coluna A vazia = fim dos dados
            break
        
        path = row[0]  # Coluna A: Path | ID
        
        # Extrair sensor_id do path (último componente)
        # Ex: Buzios\FPAB\Sensores\10S\ZN-20\10S_FD\FD-6225-2001 -> FD-6225-2001
        if path:
            path_str = str(path).strip()
            # Split por barra (pode vir com \ ou /)
            parts = path_str.replace("/", "\\").split("\\")
            sensor_id = parts[-1] if parts else None
            
            if sensor_id:
                sensores.add(sensor_id)
                linhas += 1
    
    print(f"[OK] Extraídos {len(sensores)} sensores únicos da planilha ({linhas} linhas)")
    print(f"[!] Nota: Planilha tem {linhas - len(sensores)} entradas duplicadas")
    
    # Mostrar amostra
    sample = sorted(list(sensores))[:5]
    print(f"[*] Amostra: {sample}")
    
    return sensores

def cleanup_database(sensores_xlsx, verify_only=False):
    """Remove do banco todos os sensores que não estão na planilha"""
    
    db_path = Path(__file__).parent.parent / "backend" / "safeplan.db"
    engine = create_engine(f"sqlite:///{db_path}", echo=False)
    Session = sessionmaker(bind=engine)
    session = Session()
    
    print(f"\n[*] Conectando ao banco: {db_path}")
    
    # Contar total antes
    total_before = session.query(SensorConfig).count()
    print(f"[*] Total de sensores antes: {total_before}")
    
    # Buscar sensores que estão no banco mas NÃO estão na planilha
    all_sensors = session.query(SensorConfig.id, SensorConfig.name).all()
    
    to_delete = []
    for sensor_id, name in all_sensors:
        # Extrair o ID do name (formato: "FD-6225-2001 (...)" ou "FD-6225-2001")
        if name:
            # Pegar apenas o primeira parte antes de parenteses
            name_part = name.split(" (")[0] if " (" in name else name
            
            if name_part not in sensores_xlsx:
                to_delete.append(sensor_id)
    
    print(f"\n[!] Sensores para deletar: {len(to_delete)}")
    print(f"[!] Sensores a manter: {len(all_sensors) - len(to_delete)}")
    
    if verify_only:
        print("\n[*] MODO VERIFY-ONLY: Nenhum dado foi deletado")
        
        # Mostrar amostra de sensores que seriam deletados
        if to_delete:
            sample_ids = to_delete[:5]
            sample_deletar = session.query(SensorConfig.name).filter(SensorConfig.id.in_(sample_ids)).all()
            print(f"\n[*] Amostra de sensores que SERIAM deletados:")
            for (name,) in sample_deletar:
                print(f"    - {name}")
        
        session.close()
        return False
    
    # DELETAR
    if len(to_delete) > 0:
        print(f"\n[*] Deletando {len(to_delete)} sensores...")
        
        # Deletar em lotes de 1000
        batch_size = 1000
        for i in range(0, len(to_delete), batch_size):
            batch = to_delete[i:i+batch_size]
            session.query(SensorConfig).filter(SensorConfig.id.in_(batch)).delete(synchronize_session=False)
            session.commit()
            print(f"    [*] {min(i + batch_size, len(to_delete))}/{len(to_delete)} deletados...")
    
    # Contar total depois
    total_after = session.query(SensorConfig).count()
    
    print(f"\n[RESULTADO]")
    print(f"  Antes: {total_before} sensores")
    print(f"  Depois: {total_after} sensores")
    print(f"  Deletados: {len(to_delete)}")
    
    session.close()
    return True

def main():
    verify_only = "--verify-only" in sys.argv
    
    if verify_only:
        print("\n[*] MODO: VERIFY-ONLY (sem deletar dados)")
    else:
        print("\n[!] MODO: Executar limpeza completa")
    
    try:
        # Extrair sensores da planilha
        sensores_xlsx = extract_sensors_from_xlsx()
        
        # Limpar banco
        cleanup_database(sensores_xlsx, verify_only=verify_only)
        
        if not verify_only:
            print("\n[OK] Limpeza concluída com sucesso!")
            print(f"[OK] Banco mantém apenas os {len(sensores_xlsx)} sensores da planilha Sensores.xlsx")
        else:
            print("\n[OK] Verificação concluída. Use sem --verify-only para executar a limpeza.")
        
    except Exception as e:
        print(f"\n[ERRO] {str(e)}", file=sys.stderr)
        import traceback
        traceback.print_exc()
        sys.exit(1)

if __name__ == "__main__":
    main()
