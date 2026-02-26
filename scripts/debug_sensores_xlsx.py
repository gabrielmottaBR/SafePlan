#!/usr/bin/env python3
"""Debug análise de sensores na planilha"""

from openpyxl import load_workbook
from pathlib import Path
from collections import Counter

excel_file = Path('docs/Sensores.xlsx')
wb = load_workbook(excel_file, data_only=True)
ws = wb.active

sensores = []

# Ler com row numbers
for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
    if row[0] is None:
        break
    
    path = row[0]
    
    if path:
        path_str = str(path).strip()
        parts = path_str.replace('/', '\\').split('\\')
        sensor_id = parts[-1] if parts else None
        
        if sensor_id:
            sensores.append((row_idx, sensor_id, path_str))

# Achar 'sensor_id = 2'
print('[*] Procurando por sensor_id == "2":')
found_2 = False
for row_idx, sensor_id, path in sensores:
    if sensor_id == '2':
        print(f'Linha {row_idx}: sensor_id="{sensor_id}"')
        print(f'Path: {path}')
        found_2 = True

if not found_2:
    print('  Não encontrado')

print()

# Achar duplicatas
counter = Counter([s[1] for s in sensores])
duplicatas = {k: v for k, v in counter.items() if v > 1}
print(f'[*] Total de duplicatas: {len(duplicatas)}')
if duplicatas:
    for sensor_id, count in list(duplicatas.items())[:5]:
        matching = [(row, path) for row, sid, path in sensores if sid == sensor_id]
        print(f'[*] "{sensor_id}" aparece {count} vezes:')
        for row, path in matching[:2]:
            print(f'     Linha {row}: {path}')

print()
print(f'[*] RESUMO: {len(sensores)} linhas, {len(set([s[1] for s in sensores]))} IDs únicos')
