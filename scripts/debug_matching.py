#!/usr/bin/env python3
"""Debug da lógica de matching entre planilha e banco"""

import sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).parent.parent / "backend"))

from sqlalchemy import create_engine
from sqlalchemy.orm import sessionmaker
from openpyxl import load_workbook
from src.data.models import SensorConfig

# Extrair sensores da planilha
excel_file = Path('docs/Sensores.xlsx')
wb = load_workbook(excel_file, data_only=True)
ws = wb.active

sensores_xlsx = set()

for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
    if row[0] is None:
        break
    
    path = row[0]
    
    if path:
        path_str = str(path).strip()
        parts = path_str.replace('/', '\\').split('\\')
        sensor_id = parts[-1] if parts else None
        
        if sensor_id:
            sensores_xlsx.add(sensor_id)

print(f'[*] Planilha: {len(sensores_xlsx)} sensores únicos')
print(f'[*] Amostra: {sorted(list(sensores_xlsx))[:5]}')

# Verificar banco
db_path = Path('backend/safeplan.db')
engine = create_engine(f'sqlite:///{db_path}', echo=False)
Session = sessionmaker(bind=engine)
session = Session()

all_sensors = session.query(SensorConfig.id, SensorConfig.name).all()

print(f'\n[*] Banco: {len(all_sensors)} sensores')
print(f'[*] Amostra de names: ')

# Achar exemplos que match e não match
match_count = 0
not_match_count = 0
not_match_sample = []

for sensor_id, name in all_sensors:
    if name:
        # Extrair o primeiro componente do name (antes do parentese)
        name_part = name.split(' (')[0] if ' (' in name else name
        
        if name_part in sensores_xlsx:
            match_count += 1
        else:
            not_match_count += 1
            if len(not_match_sample) < 10:
                not_match_sample.append((sensor_id, name, name_part))

print(f'[*] Em TODOS os {len(all_sensors)} sensores do banco:')
print(f'    - Match com planilha: {match_count}')
print(f'    - Não match: {not_match_count}')

print(f'\n[*] Exemplos de NON-MATCH (sensores a deletar):')
for sid, name, part in not_match_sample:
    print(f'    ID={sid}, Name={name}')
    print(f'      Tried to match: "{part}"')
    print()

session.close()
