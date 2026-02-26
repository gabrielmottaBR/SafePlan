#!/usr/bin/env python3
"""
Script para limpar completamente o banco e importar APENAS os 9.983 sensores da planilha Sensores.xlsx

Uso:
    python scripts/reset_db_from_sensores_xlsx.py [--verify-only]
"""

import sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).parent.parent / "backend"))

from sqlalchemy import create_engine, text
from sqlalchemy.orm import sessionmaker
from openpyxl import load_workbook
from src.data.models import SensorConfig

def limpar_banco():
    """Deleta TODOS os registros do banco"""
    
    db_path = Path(__file__).parent.parent / "backend" / "safeplan.db"
    engine = create_engine(f"sqlite:///{db_path}", echo=False)
    Session = sessionmaker(bind=engine)
    session = Session()
    
    # Contar antes
    total_antes = session.query(SensorConfig).count()
    print(f"\n[*] Total de sensores antes: {total_antes}")
    
    # Deletar todos
    print(f"[*] Deletando TODOS os sensores...")
    session.query(SensorConfig).delete(synchronize_session=False)
    session.commit()
    
    # Contar depois
    total_depois = session.query(SensorConfig).count()
    print(f"[*] Total de sensores depois: {total_depois}")
    
    session.close()
    return engine

def importar_sensores_xlsx(engine):
    """Importa os 9.983 sensores da planilha Sensores.xlsx"""
    
    excel_file = Path(__file__).parent.parent / "docs" / "Sensores.xlsx"
    
    if not excel_file.exists():
        raise FileNotFoundError(f"Arquivo não encontrado: {excel_file}")
    
    print(f"\n[*] Lendo planilha: {excel_file}")
    
    wb = load_workbook(excel_file, data_only=True)
    ws = wb.active
    
    Session = sessionmaker(bind=engine)
    session = Session()
    
    sensores_adicionados = 0
    
    # Pular header (linha 1), ler todas as linhas
    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if row[0] is None:
            break
        
        path = row[0]  # Coluna A: Path | ID (Full AF path)
        grupo = row[2] if len(row) > 2 else None  # Coluna C: Grupo
        modulo = row[3] if len(row) > 3 else None  # Coluna D: Módulo
        
        if path:
            path_str = str(path).strip()
            
            # Extrair sensor_id do path (último componente)
            parts = path_str.replace("/", "\\").split("\\")
            sensor_id = parts[-1] if parts else None
            
            # Extrair grupo do path (penúltimo componente)
            grupo_from_path = parts[-2] if len(parts) > 1 else None
            
            # Extrair modulo do path (3º-último componente)
            modulo_from_path = parts[-3] if len(parts) > 2 else None
            
            if sensor_id:
                # Criar registro
                sensor = SensorConfig(
                    sensor_id=sensor_id,
                    name=sensor_id,  # Nome é o ID do sensor
                    sensor_type="UNKNOWN",
                    location="Buzios",
                    unit="ppm",
                    grupo=grupo_from_path if grupo is None else grupo,
                    modulo=modulo_from_path if modulo is None else modulo,
                    is_active=True
                )
                
                session.add(sensor)
                sensores_adicionados += 1
                
                # Commit a cada 1000 sensores
                if sensores_adicionados % 1000 == 0:
                    session.commit()
                    print(f"  [*] {sensores_adicionados} sensores processados...")
    
    # Commit final
    session.commit()
    
    print(f"\n[OK] {sensores_adicionados} sensores importados da planilha")
    
    session.close()
    return sensores_adicionados

def validar_resultado(engine):
    """Valida o resultado final"""
    
    Session = sessionmaker(bind=engine)
    session = Session()
    
    total = session.query(SensorConfig).count()
    com_grupo = session.query(SensorConfig).filter(SensorConfig.grupo != None).count()
    com_modulo = session.query(SensorConfig).filter(SensorConfig.modulo != None).count()
    
    print(f"\n[VALIDAÇÃO]")
    print(f"  Total sensores: {total}")
    print(f"  Com Grupo: {com_grupo}")
    print(f"  Com Módulo: {com_modulo}")
    
    if total == 9983:
        print(f"\n[OK] ✅ Banco contém EXATAMENTE 9.983 sensores!")
        return True
    else:
        print(f"\n[ERRO] ❌ Banco deveria ter 9.983 sensores, mas tem {total}")
        return False
    
    session.close()

def main():
    print("\n" + "="*70)
    print("  RESET COMPLETO DO BANCO - IMPORTAR SENSORES.XLSX (9.983 sensores)")
    print("="*70)
    
    verify_only = "--verify-only" in sys.argv
    
    if verify_only:
        print("\n[*] MODO: VERIFY-ONLY (sem fazer mudanças)")
    else:
        print("\n[!] MODO: Limpeza completa e re-importação")
    
    try:
        # Limpar banco
        if not verify_only:
            engine = limpar_banco()
        else:
            db_path = Path(__file__).parent.parent / "backend" / "safeplan.db"
            engine = create_engine(f"sqlite:///{db_path}", echo=False)
            Session = sessionmaker(bind=engine)
            session = Session()
            total = session.query(SensorConfig).count()
            print(f"\n[*] Modo VERIFY: Banco tem {total} sensores")
            session.close()
            print("[*] Para executar: Remove --verify-only")
            return
        
        # Importar sensores
        sensores_importados = importar_sensores_xlsx(engine)
        
        # Validar
        validar_resultado(engine)
        
    except Exception as e:
        print(f"\n[ERRO] {str(e)}", file=sys.stderr)
        import traceback
        traceback.print_exc()
        sys.exit(1)

if __name__ == "__main__":
    main()
