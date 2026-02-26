#!/usr/bin/env python3
"""
Script para importar sensores do arquivo docs/Sensores.xlsx
para o banco de dados SafePlan com Grupo e Módulo.

Uso:
    python scripts/import_sensores_xlsx.py [--limit=N] [--verify-only]
"""
import sys
import os
from pathlib import Path

# Add project root to path
project_root = Path(__file__).parent.parent
sys.path.insert(0, str(project_root))

from openpyxl import load_workbook
from sqlalchemy import create_engine
from sqlalchemy.orm import sessionmaker
from backend.src.data.models import Base, SensorConfig
from backend.config.settings import Settings

# For data_only, we need the calculated values
# The Excel file has formulas in columns C & D that extract from column A paths


def extract_sensor_id_from_path(path: str) -> str:
    """Extract sensor ID from AF path like 'Buzios\\FPAB\\Sensores\\10S\\ZN-20\\10S_FD\\FD-6225-2001'"""
    if not path:
        return "Unknown"
    # The sensor ID is the last part of the path
    parts = str(path).split('\\')
    return parts[-1] if parts else "Unknown"


def parse_grupo_from_path(path: str) -> str:
    """Extract Grupo (voting group) from AF path"""
    if not path:
        return None
    # Grupo is typically like "10S_FD" from the path
    # Try to extract from the path structure
    parts = str(path).split('\\')
    # Usually penultimate part like "10S_FD"
    if len(parts) >= 2:
        return parts[-2] if parts[-2] else None
    return None


def parse_modulo_from_path(path: str) -> str:
    """Extract Módulo from AF path"""
    if not path:
        return None
    # Módulo is typically the first numeric part like "10S"
    parts = str(path).split('\\')
    # Usually third-last part like "10S"
    if len(parts) >= 3:
        return parts[-3] if parts[-3] else None
    return None


def import_sensores(limit=None, verify_only=False):
    """Import sensors from Excel to database."""
    
    print("\n" + "="*80)
    print(f"SafePlan - Importação de Sensores a partir de Excel")
    print("="*80 + "\n")
    
    # Load Excel file
    print("[1/4] Carregando planilha...")
    excel_file = project_root / 'docs' / 'Sensores.xlsx'
    
    if not excel_file.exists():
        print(f"[ERROR] Arquivo não encontrado: {excel_file}")
        return False
    
    try:
        wb = load_workbook(excel_file, data_only=True)
        ws = wb[wb.sheetnames[0]]
        
        total_rows = ws.max_row - 1  # Exclude header
        if limit:
            total_rows = min(total_rows, limit)
        
        print(f"[OK] Planilha carregada: {total_rows} sensores encontrados")
        print(f"   Sheet: {ws.title}")
        print()
        
        if verify_only:
            print("[*] Modo VERIFY-ONLY: apenas validando dados\n")
        
        # Parse header
        headers = {}
        for col in range(1, ws.max_column + 1):
            headers[col] = ws.cell(1, col).value
        
        print(f"[2/4] Colunas encontradas:")
        for col, header in headers.items():
            print(f"       {chr(64+col)}: {header}")
        print()
        
        # Initialize database if not verify-only
        if not verify_only:
            print("[3/4] Inicializando banco de dados...")
            settings = Settings()
            engine = create_engine(settings.database_url, echo=False)
            
            # Create tables
            Base.metadata.create_all(bind=engine)
            Session = sessionmaker(bind=engine)
            session = Session()
            print("[OK] Banco inicializado\n")
        else:
            print("[3/4] Modo verify - banco não será modificado\n")
        
        # Import sensors
        print("[4/4] Processando sensores...")
        print("-" * 80)
        
        imported = 0
        skipped = 0
        errors = 0
        
        for row_num in range(2, 2 + total_rows):
            try:
                # Get values
                path_id = ws.cell(row_num, 1).value
                path_grupo = ws.cell(row_num, 2).value
                grupo_col_c = ws.cell(row_num, 3).value
                modulo_col_d = ws.cell(row_num, 4).value
                
                # Use the Excel formulas if available, otherwise parse from path
                path_to_use = path_id or path_grupo or ""
                grupo = grupo_col_c if grupo_col_c else parse_grupo_from_path(path_to_use)
                modulo = modulo_col_d if modulo_col_d else parse_modulo_from_path(path_to_use)
                sensor_id = extract_sensor_id_from_path(path_to_use)
                
                # Skip if no sensor ID
                if not sensor_id or sensor_id == "Unknown":
                    skipped += 1
                    continue
                
                if verify_only:
                    # Just show what would be imported
                    if imported < 3:  # Show first 3
                        print(f"Row {row_num}:")
                        print(f"  Sensor ID: {sensor_id}")
                        print(f"  Grupo: {grupo}")
                        print(f"  Módulo: {modulo}")
                        print(f"  Path: {str(path_to_use)[:60]}")
                        print()
                    imported += 1
                else:
                    # Check if sensor already exists
                    existing = session.query(SensorConfig).filter(
                        SensorConfig.sensor_id == sensor_id
                    ).first()
                    
                    if existing:
                        # Update existing sensor with grupo/modulo
                        existing.grupo = grupo
                        existing.modulo = modulo
                        session.commit()
                    else:
                        # Create new sensor
                        sensor = SensorConfig(
                            sensor_id=sensor_id,
                            name=sensor_id,  # Use sensor_id as name for now
                            sensor_type="GAS_DETECTOR",  # Default type
                            location="Buzios",  # Default location
                            unit="ppm",  # Default unit
                            grupo=grupo,
                            modulo=modulo,
                            is_active=True,
                        )
                        session.add(sensor)
                    
                    imported += 1
                    
                    if imported % 1000 == 0:
                        if not verify_only:
                            session.commit()
                        print(f"  [*] {imported} sensores processados...")
                
            except Exception as e:
                print(f"  [ERROR] Row {row_num}: {e}")
                errors += 1
                continue
        
        # Final commit
        if not verify_only:
            session.commit()
            session.close()
        
        print("-" * 80)
        print()
        print("[RESULTADO]")
        print(f"  Importados/Validados: {imported}")
        if errors:
            print(f"  Erros: {errors}")
        if skipped:
            print(f"  Pulados (sem sensor_id): {skipped}")
        print()
        
        if verify_only:
            print("[OK] Validação concluída. Para importar, remova --verify-only")
        else:
            print(f"[OK] {imported} sensores importados com Grupo e Módulo")
        
        return True
        
    except Exception as e:
        print(f"[FATAL ERROR] {e}")
        import traceback
        traceback.print_exc()
        return False


if __name__ == '__main__':
    import argparse
    
    parser = argparse.ArgumentParser(description='Importar sensores de Sensores.xlsx')
    parser.add_argument('--limit', type=int, help='Limite de sensores a importar')
    parser.add_argument('--verify-only', action='store_true', help='Apenas verificar sem importar')
    
    args = parser.parse_args()
    
    success = import_sensores(limit=args.limit, verify_only=args.verify_only)
    sys.exit(0 if success else 1)
